import pandas as pd
import requests
import time
import xml.etree.ElementTree as ET
import os
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from urllib.parse import urlparse
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

session = requests.Session()
retry_strategy = Retry(
    total=3,
    backoff_factor=0.5,
    status_forcelist=[429, 500, 502, 503, 504],
)

adapter = HTTPAdapter(max_retries=retry_strategy)
session.mount("http://", adapter)
session.mount("https://", adapter)

# -----------------------------
# CONFIG
# -----------------------------
# IRS CSV URLs (Regions 1-4)
CSV_DIR = "CSVs"

OUTPUT_FILE = "output_utf8.csv"

# XML fields we want to extract
FIELDS = [
    "CYTotalRevenueAmt",
    "CYContributionsGrantsAmt",
    "MembershipDuesAmt",
    "GovernmentGrantsAmt",
    "NoncashContributionsAmt",
    "TotalVolunteersCnt"
]

# Delay between requests (seconds)
REQUEST_DELAY = 0.5


# -----------------------------
# FUNCTION: Download IRS CSV
# -----------------------------
def download_irs_csv(url, filename):
    try:
        print(f"Downloading {filename} from IRS...")
        res = session.get(url, timeout=30)
        res.raise_for_status()
        with open(filename, "wb") as f:
            f.write(res.content)
        print(f"✅ Saved {filename}")
    except Exception as e:
        print(f"[ERROR] Could not download {filename}: {e}")


# -----------------------------
# FUNCTION: Get ProPublica XML URL
# -----------------------------
def get_xml_url(ein):
    """
    Retrieve the XML link from ProPublica for a given EIN.
    """
    base_url = f"https://projects.propublica.org/nonprofits/organizations/{ein}"
    try:
        res = session.get(base_url, timeout=10)
        if res.status_code != 200:
            return None

        # Search for XML link
        soup = BeautifulSoup(res.text, "html.parser")
        tag = soup.find("a", string="XML")
        if tag:
            hreflink = tag.get("href");
            return f"https://projects.propublica.org/{hreflink}"

        #match = re.search(r'https://[^\s"]+\.xml', res.text)
        #if match:
            #return match.group(0)
    except Exception as e:
        print(f"[ERROR] EIN {ein}: {e}")
    return None


# -----------------------------
# FUNCTION: Extract XML Fields
# -----------------------------
def extract_xml_data(xml_url):
    """
    Download XML and extract specified fields.
    """
    data = {field: None for field in FIELDS}
    try:
        res = session.get(xml_url, timeout=10)
        res.raise_for_status()
        root = ET.fromstring(res.content)

        for elem in root.iter():
            tag = elem.tag.split("}")[-1]  # remove namespace
            if tag in FIELDS:
                data[tag] = elem.text
    except Exception as e:
        print(f"[XML ERROR] {xml_url}: {e}")
    return data


# -----------------------------
# FUNCTION: Map File Url from url.xlsx
# -----------------------------
def build_irs_file_map(excel_path, output_dir ):
    """
    Reads an Excel file containing URLs and returns a dictionary mapping
    local file paths to their corresponding URLs.
    """
    os.makedirs(output_dir, exist_ok=True)
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    file_map = {}
    for row in sheet.iter_rows(values_only=True):
        url = row[0]
        if not url:
            continue  # skip empty cells
        filename = os.path.basename(urlparse(url).path)
        local_path = os.path.join(output_dir, filename)
        file_map[local_path] = url

    return file_map


# Usage
# -----------------------------
# FUNCTION: Process with ein Pipeline
# -----------------------------
def process_ein(ein):
    if not ein:
        return {field: None for field in FIELDS}

    try:
        print(f"Processing EIN: {ein}")

        xml_url = get_xml_url(ein)
        if not xml_url:
            return {field: None for field in FIELDS}

        data = extract_xml_data(xml_url)
        print(data)
        return data

    except Exception as e:
        print(f"[ERROR] EIN {ein}: {e}")
        return {field: None for field in FIELDS}

# -----------------------------
# FUNCTION: Main Pipeline
# -----------------------------
def process_data():
    # Step 1: Download all IRS CSVs    
    os.makedirs(CSV_DIR, exist_ok=True)

    IRS_FILES = build_irs_file_map("url.xlsx", CSV_DIR)

    for local_file, url in IRS_FILES.items():
        download_irs_csv(url, local_file)

    # Step 2: Read and combine CSVs
    all_dfs = []
    for local_file in IRS_FILES.keys():
        df = pd.read_csv(local_file, dtype={"EIN": str}, encoding="latin-1")
        all_dfs.append(df)
    df_combined = pd.concat(all_dfs, ignore_index=True)
    print(f"✅ Combined {len(df_combined)} rows from all regions")

    # Step 3: Process each EIN
    eins = df_combined["EIN"].tolist()
    results = [None] * len(eins)

    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {
            executor.submit(process_ein, ein): i
            for i, ein in enumerate(eins)
        }

        for future in as_completed(futures):
            index = futures[future]
            results[index] = future.result()

    # Step 4: Merge results and save
    result_df = pd.DataFrame(results)
    final_df = pd.concat([df_combined, result_df], axis=1)
    final_df.to_csv(OUTPUT_FILE, index=False, encoding="utf-8")
    print(f"✅ Done! Output saved to {OUTPUT_FILE}")


# -----------------------------
# RUN SCRIPT
# -----------------------------
if __name__ == "__main__":
    process_data()
